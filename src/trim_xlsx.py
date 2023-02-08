# Roughly trim the incoming XLSX documents to process them faster in the next step

# Remove all empty rows > 30
# Remove all excess columns
import os
from pathlib import Path

from openpyxl import load_workbook
from lib.utils import get_input_files


def get_first_empty_cell(seq, is_row):
    for i, v in enumerate(seq):
        if not v.value:
            return i + 1
    # Reached the end! No empty cells
    return None


def trim_file(path):
    print(path)
    wb = load_workbook(path)
    for ws in wb:
        if ws["A1"].value is None:
            ws["A1"] = "NULL"

        header = ws["1"]
        col_del = get_first_empty_cell(header, True)
        if col_del:
            width = len(header)
            col_dist = width - col_del + 1
            ws.delete_cols(col_del, col_dist)

        margin = ws["A"]
        row_del = get_first_empty_cell(margin, False)
        if row_del:
            height = len(margin)
            row_dist = height - row_del + 1
            ws.delete_rows(row_del, row_dist)

        output_path = os.path.join("output/xlsx", Path(path).name)
    wb.save(output_path)


if __name__ == "__main__":
    input_files = get_input_files("input/xlsx")
    for file in input_files:
        trim_file(file)
