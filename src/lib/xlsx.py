from openpyxl import Workbook, load_workbook


class ExcelFile:
    def __init__(self, file_path=None):
        if file_path:
            self.wb = load_workbook(file_path)
        else:
            self.wb = Workbook()

    def get(self, sheet_name: str, cell: str, fallback="NULL"):
        try:
            return self.wb[sheet_name][cell].value
        except KeyError:
            return fallback

    def gets(self, sheet_name: str, cells: str, fallback):
        try:
            return [cell.value for cell in self.wb[sheet_name][cells]]
        except KeyError:
            return fallback

    def set(self, sheet_name: str, cell: str, value):
        self.wb[sheet_name][cell] = value
